# @File: Create_Template.py
# @Time: 2024/1/28 下午 10:32
# @Author: Nan1_Chen
# @Mail: Nan1_Chen@pegatroncorp.com

import re
import openpyxl
from json import loads
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Alignment
from LoadFilePath import data_pardir, export_dir
from ExtractZIP import file_name
from MarkPoint import get_alert_list


def create_template():
    alert_list = get_alert_list()
    wb = openpyxl.Workbook()
    ws = wb.active
    print(type(ws))

    # 格式設置
    ws.row_dimensions[1].height = 40
    fill_head = PatternFill('solid', fgColor='adaaa6')  # 设置填充颜色为 灰色
    fill_ng = PatternFill('solid', fgColor='ff99cc')  # 设置填充颜色为 紅色
    fill_ok = PatternFill('solid', fgColor='aacf91')  # 设置填充颜色为 綠色
    fill_error = PatternFill('solid', fgColor='ffcf71')  # 设置填充颜色为 橙色
    font = Font(u'Calibri', size=12, color='ffffff', bold=True,
                italic=False, strike=False)  # 设置字体样式
    ws['a1'].fill = fill_head  # 应用填充样式在A1单元格
    ws['b1'].fill = fill_head  # 应用填充样式在B1单元格
    ws['c1'].fill = fill_head  # 应用填充样式在C1单元格
    ws['d1'].fill = fill_head  # 应用填充样式在D1单元格
    ws['e1'].fill = fill_head  # 应用填充样式在E1单元格
    ws['f1'].fill = fill_head  # 应用填充样式在F1单元格
    ws['g1'].fill = fill_head  # 应用填充样式在G1单元格
    ws['h1'].fill = fill_head  # 应用填充样式在H1单元格
    ws['i1'].fill = fill_head  # 应用填充样式在I1单元格
    ws['j1'].fill = fill_head  # 应用填充样式在J1单元格
    ws['k1'].fill = fill_head  # 应用填充样式在J1单元格
    ws['a1'].font = font  # 应用填充样式在A1单元格
    ws['b1'].font = font  # 应用填充样式在B1单元格
    ws['c1'].font = font  # 应用填充样式在C1单元格
    ws['d1'].font = font  # 应用填充样式在D1单元格
    ws['e1'].font = font  # 应用填充样式在E1单元格
    ws['f1'].font = font  # 应用填充样式在F1单元格
    ws['g1'].font = font  # 应用填充样式在G1单元格
    ws['h1'].font = font  # 应用填充样式在H1单元格
    ws['i1'].font = font  # 应用填充样式在I1单元格
    ws['j1'].font = font  # 应用填充样式在J1单元格
    ws['k1'].font = font  # 应用填充样式在J1单元格
    # 向excel中写入表头
    ws['a1'] = 'SN'
    ws['b1'] = 'Station ID'
    ws['c1'] = 'Test Time'
    ws['d1'] = 'BGI Image'
    ws['e1'] = 'ICE Image'
    ws['f1'] = 'SOB Image'
    ws['g1'] = 'BGI Issue'
    ws['h1'] = 'ICE Issue'
    ws['i1'] = 'SOB Issue'
    ws['j1'] = 'Root Cause'
    ws['k1'] = 'CA'
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 30

    # 插入信息：
    x = 2  # 在第一行开始写
    y = 1  # 在第一列开始写
    kk = 1

    for unit_folder_path in data_pardir.iterdir():
        print(unit_folder_path)
        sn = unit_folder_path.name.split('_')[0]
        # print(SN)
        print(
            f"*****第 {kk} 個機臺開始插入信息************************************************************")
        # 1.插入第一欄SN信息：
        print(f"-----1.正在插入A欄 Serial Number 信息: {sn}---------")
        ws.cell(row=x, column=y).value = sn
        if unit_folder_path.name in alert_list:
            ml_info = alert_list[unit_folder_path.name]
            station_id = ml_info.get('station_id')
            test_time = ml_info.get('uut_start')
            # 2.插入第二欄Station ID信息：
            print(
                f"-----2.正在插入B欄 Station ID 信息: {station_id}----------------")
            # Station ID
            ws.cell(row=x, column=y + 1).value = station_id
            # 3.插入第三欄Test Time信息：
            print(
                f"-----3.正在插入C欄 Test Time 信息: {test_time} -------------")
            ws.cell(row=x, column=y + 2).value = test_time  # Time

        # 4.插入第四欄BGI Image信息：
        print('-----4.正在插入D欄 BGI圖片--------------------')

        for pic_path in file_name(unit_folder_path, r'^(SOBBK|ICEBK|BGI|SOB|ICE).*\.JPG$'):
            print(pic_path.stem)
            image = str(pic_path)  # 返回文件名
            image_type = re.findall(
                r'^(SOBBK|ICEBK|BGI|SOB|ICE).*\.JPG$', pic_path.name)[0]
            im = PILImage.open(pic_path)
            w, h = im.size
            im.thumbnail((w // 3, h // 3))
            im.save(pic_path, 'JPEG')
            img_column = {'SOB': 'F', 'ICE': 'E', 'SOBBK': 'F',
                          'ICEBK': 'E', 'BGI': 'D'}.get(image_type, 'D')

            # 获取图片
            img = Image(image)
            img.anchor = img_column + str(x)
            # 设置图片的大小
            img.width, img.height = (100, 100)
            # 设置表格的宽20和高85
            ws.column_dimensions[img_column].width = 20
            ws.row_dimensions[x].height = 85
            # 图片插入名称对应单元格
            ws.add_image(img)

        # 7.插入第七欄BGI Issue信息：
        print('-----7.正在插入H欄 BGI Ml infor--------------------')
        for text_path in file_name(unit_folder_path, r'.*(SOBBK|ICEBK|BGI)\.txt'):
            text = str(text_path)  # 返回文件名
            text_type = re.findall(
                r'.*(SOBBK|ICEBK|BGI|SOB|ICE)\.txt', text_path.name)[0]
            y_offset = {'SOB': 8, 'ICE': 7, 'SOBBK': 8,
                        'ICEBK': 7, 'BGI': 6}.get(text_type, '6')
            with open(text, "r", encoding='utf-8') as f:
                content = f.read()
            result = re.findall(r'^\{.*}$', content, re.M)
            if result:
                result_info = loads(result[0])
                ml_status = result_info.get('decision', 999)

                if ml_status == 0:
                    # print("該機臺bgi無異常")
                    ws.cell(row=x, column=y + y_offset).value = "no issue"
                    ws.cell(row=x, column=y + y_offset).fill = fill_ok
                else:
                    result1 = ";\n".join(result_info.get('issues', []))
                    ws.cell(row=x, column=y + y_offset).value = result1
                    ws.cell(row=x, column=y + y_offset).fill = fill_ng
            else:
                ws.cell(row=x, column=y + y_offset).value = 'not found response body'
                ws.cell(row=x, column=y + y_offset).fill = fill_error
        print(
            f"*****第 {kk} 個機臺信息插入完成************************************************************")
        print("                                                     ")
        x += 1
        kk += 1

    max_rows = ws.max_row  # 获取最大行
    max_columns = ws.max_column  # 获取最大列
    align = Alignment(horizontal='center', vertical='center', wrapText=True)
    # openpyxl的下标从1开始
    for iiii in range(1, max_rows + 1):
        for jjjj in range(1, max_columns + 1):
            ws.cell(iiii, jjjj).alignment = align
    from LoadFilePath import ml_station
    excel_path = export_dir / f'{ml_station} ML Alert SN re-judge tracker list.xlsx'
    wb.save(excel_path)
    wb.close()


if __name__ == "__main__":
    create_template()
