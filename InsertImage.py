import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles.borders import (
    BORDER_MEDIUM,
    BORDER_DOUBLE,
    BORDER_THIN,
    BORDER_THICK,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from LoadFilePath import export_dir

# 表头填充样式,背景颜色#8db4e2
fill_head = PatternFill(
    "solid",
    fgColor="8db4e2",
)
# 表头文本样式,Calibri 28号字体,颜色#000000,加粗
font_head = Font(
    "Calibri",
    size=28,
    color="000000",
    bold=True,
    italic=False,
    strike=False,
)
# 内容文本样式,Calibri 24号字体,颜色#000000
font_body = Font(
    "Calibri",
    size=24,
    color="000000",
    bold=False,
    italic=False,
    strike=False,
)
# 居中对齐,自动换行
align = Alignment(
    horizontal="center",
    vertical="center",
    text_rotation=0,
    wrapText=True,
    shrink_to_fit=False,
    indent=0,
)


class MLAlertTable:
    def __init__(self, test_info: DataFrame):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.worksheets[0]
        self.worksheet.title = "Sheet1"
        self.test_info = test_info
        self.machine_name = '&'.join(test_info['display_name'].unique())
        self.output_path = export_dir / f'{self.machine_name} ML Alert SN re-judge tracker list.xlsx'
        self.headers = []
        self.fill_data_frame()
        self.max_row = self.worksheet.max_row
        self.max_column = self.worksheet.max_column
        self.set_header()
        self.set_header_border()
        self.set_body()
        self.set_body_border()
        self.export()

    def fill_data_frame(self):
        for row in dataframe_to_rows(self.test_info, index=False, header=True):
            self.worksheet.append(row)

    def set_header(self):
        self.worksheet.row_dimensions[1].height = 80  # type: ignore

        for i in range(self.max_column):
            col_letter = get_column_letter(i + 1)
            ccw = self.worksheet.column_dimensions[col_letter].width
            head = self.worksheet[f"{col_letter}1"]
            head.fill = fill_head
            head.font = font_head
            head.alignment = align

            if ccw < 27 and head.value.endswith("_pic"):
                self.worksheet.column_dimensions[col_letter].width = 54
            elif ccw < 20:
                self.worksheet.column_dimensions[col_letter].width = 40
            self.headers.append(head.value)

    def set_body(self):
        for j in range(2, self.max_row + 1):
            self.worksheet.row_dimensions[j].height = 170  # type: ignore
            for i in range(1, self.max_column + 1):
                head = self.headers[i - 1]
                cell = self.worksheet.cell(j, i)
                cell.alignment = align
                cell.font = font_body
                if head.endswith("_pic") and cell.value != "NA":
                    self.insert_image_to_cell(j - 1, i - 1, cell.value)
                    cell.value = ""

    def get_cell_width_height(self, row, col) -> tuple:
        col_letter = get_column_letter(col + 1)
        ccw = self.worksheet.column_dimensions[col_letter].width
        crh = self.worksheet.row_dimensions[row + 1].height
        w = (ccw * 72) // 9
        h = (crh * 18) // 13.5
        r = w / h
        return w, h, r

    def insert_image_to_cell(self, row, col, image_url):
        img = Image(image_url)
        _from = AnchorMarker(col, 100000, row, 100000)
        _to = AnchorMarker(col + 1, -100000, row + 1, -100000)
        img.anchor = TwoCellAnchor("twoCell", _from, _to)  # type: ignore
        self.worksheet.add_image(img)

    @staticmethod
    def set_border(
        t_border=BORDER_THIN,
        b_border=BORDER_THIN,
        l_border=BORDER_THIN,
        r_border=BORDER_THIN,
        t_color="000000",
        b_color="000000",
        l_color="000000",
        r_color="000000",
    ) -> Border:
        border = Border(
            top=Side(border_style=t_border, color=t_color),
            bottom=Side(border_style=b_border, color=b_color),
            left=Side(border_style=l_border, color=l_color),
            right=Side(border_style=r_border, color=r_color),
        )
        return border

    def set_header_border(self):
        self.worksheet.cell(1, 1).border = self.set_border(
            BORDER_THICK, BORDER_MEDIUM, BORDER_THICK, BORDER_THIN
        )
        self.worksheet.cell(1, self.max_column).border = self.set_border(
            BORDER_THICK, BORDER_MEDIUM, BORDER_THIN, BORDER_THICK
        )
        for i in range(2, self.max_column):
            self.worksheet.cell(1, i).border = self.set_border(
                BORDER_THICK, BORDER_MEDIUM, BORDER_THIN, BORDER_THIN
            )

    def set_body_border(self):
        for row in range(2, self.max_row):
            left_cell = self.worksheet.cell(row, 1)
            right_cell = self.worksheet.cell(row, self.max_column)
            left_cell.border = self.set_border(
                left_cell.border.top.style, BORDER_DOUBLE, BORDER_THICK, BORDER_THIN
            )
            right_cell.border = self.set_border(
                right_cell.border.top.style, BORDER_DOUBLE, BORDER_THIN, BORDER_THICK
            )
            for col in range(2, self.max_column):
                cell = self.worksheet.cell(row, col)
                cell.border = self.set_border(
                    cell.border.top.style, BORDER_DOUBLE, BORDER_THIN, BORDER_THIN
                )
        left_bottom_cell = self.worksheet.cell(self.max_row, 1)
        right_bottom_cell = self.worksheet.cell(self.max_row, self.max_column)
        left_bottom_cell.border = self.set_border(
            left_bottom_cell.border.top.style, BORDER_THICK, BORDER_THICK, BORDER_THIN
        )
        right_bottom_cell.border = self.set_border(
            right_bottom_cell.border.top.style, BORDER_THICK, BORDER_THIN, BORDER_THICK
        )
        for col in range(2, self.max_column):
            bottom_cell = self.worksheet.cell(self.max_row, col)
            bottom_cell.border = self.set_border(
                bottom_cell.border.top.style, BORDER_THICK, BORDER_THIN, BORDER_THIN
            )

    def export(self):
        self.workbook.save(self.output_path)


if __name__ == "__main__":
    pass
